from django.shortcuts import render
from django.http import HttpResponse
from django.forms import inlineformset_factory
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.http import Http404
from django.contrib.auth.models import Group

from .models import *
from .forms import *
from .decorators import *

from datetime import *

def error404(request, *args, **kwargs):
    return render(request,'pages/404.html')

def error500(request, *args, **kwargs):
    return render(request,'pages/500.html')
def export_movies_to_xlsx(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    student_queryset = Student.objects.all()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Students'

    # Define the titles for columns
    columns = [
        'ID',
        'Title',
        'Description',
        'Length',
        'Rating',
        'Price',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for movie in student_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            movie.pk,
            movie.fullname,
            movie.phonenumber,
            movie.classname,
            movie.create_date,
            movie.active,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
# Create your views here.
@unauthenticated_login
def loginPage(request):
        if request.method == 'POST':
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request, username=username, password=password)
            context={'user':user}
            if user is not None:
                login(request, user )
                return redirect('home')
            else:
                messages.info(request, 'Tên đăng nhập hoặc mật khẩu không đúng')
                return render(request, 'pages/login.html', context)
        context={}
        return render(request, 'pages/login.html', context)

@unauthenticated_register
def registerPage(request):
        form = CreateUserForm()
        if request.method == 'POST':
            form = CreateUserForm(request.POST)
            if form.is_valid():
                # user.is_superuser = 1
                # user.is_staff = 1
                # user = form.save()  
                username = form.cleaned_data.get('username')

                # group = Group.objects.get(name='manager')
                # user.groups.add(group)
                user = form.save()
                if str(request.POST.get('is_active')) == 'on':
                    is_active = 1
                else:
                    is_active = 0
                if str(request.POST.get('is_staff')) == 'on':
                    is_staff = 1
                else:
                    is_staff = 0
                if str(request.POST.get('is_superuser')) == 'on':
                    is_superuser = 1
                else:
                    is_superuser = 0

                if is_superuser:
                    group = Group.objects.get(name='admin')
                    user.groups.add(group)
                elif is_staff:
                    group = Group.objects.get(name='manager')
                    user.groups.add(group)
                else:
                    group = Group.objects.get(name='staff')
                    user.groups.add(group)
                messages.success(request, 'Tài khoản ' + username + ' tạo thành công')
                return redirect('login')
            else:
                messages.error(request, "Lỗi tạo tài khoản")
                return render(request, 'pages/register.html')
        context = {'form':form}
        return render(request, 'pages/register.html',context)
        # context = {'form':form}
        # return render(request, 'pages/register.html', context)

def logoutUser(request):
    logout(request)
    return redirect('login')
    
@login_required(login_url='login')
def home(request):
    notify = Notify.objects.all()
    listclass = Classname.objects.all()
    for i1 in listclass:
        startdate = datetime.strptime(str(i1.startdate), '%Y-%m-%d')
        today = datetime.now()
        student = Student.objects.all()
        if startdate <= today:
            # print(i1.pk) 
            i1.active=1
            i1.save()
        for a in student:
            if i1.pk==a.classname_id:
                if startdate <= today:
                    # print(a.classname_id) 
                    a.active=1
                    a.save()
                    
            
                # if a.classname_id==i1.pk:
                #     a.active = 1
                #     a.save()
    schedule = Schedule.objects.all()
    for i2 in schedule:
        daylearn = datetime.strptime(str(i2.daylearn), '%Y-%m-%d')
        today = datetime.now()
        if daylearn < today:
            i2.active=1
            i2.save()
    context = {'schedule':schedule,'notify':notify}
    return render(request, 'pages/home.html',context)

# Class 
# Dat
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff'])
def listClass(request):
    # auto run 
    schedule = Schedule.objects.all()
    listclass = Classname.objects.all() 
    arraya = []
    for itema in listclass:
        SLa = {}
        SLa['idclass'] = itema.pk
        SLa['number'] = Student.objects.filter(classname_id=itema).count()
        arraya.append(SLa)
    arrayb = []
    for itemb in listclass:
        SLb = {}
        SLb['class'] = itemb.fullname
        SLb['number'] = Schedule.objects.filter(classname=itemb,active=1).count()
        arrayb.append(SLb)
    context = {'listclass': listclass,'schedule':schedule,'SLa':arraya,'SLb':arrayb}
    return render(request, 'class/listclass.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager'])
def createClass(request):
    form = CreateClassnameForm()
    unit = Unit.objects.all()
    area = Area.objects.all()
    room = Room.objects.all()
    timeshift = TimeShift.objects.all()
    timeweek = TimeWeek. objects.all()
    teacher = Teacher.objects.all()
    if request.method == 'POST':
        form = CreateClassnameForm(request.POST)
        if form.is_valid():
            form.save()
            count = 0
            start_date = datetime.strptime(str(request.POST.get('startdate')), '%Y-%m-%d')
            step = timedelta(days=1)
            time = TimeShift.objects.get(pk = request.POST.get('timeshift'))
            if int(request.POST.get('timeweek')) == 2:
                day = [1,3,5]
            else:
                day = [0,2,4]
            while count < int(request.POST.get('datecount')):
                for i in day:
                    if i==start_date.date().weekday():
                        print(start_date.date())
                        Schedule.objects.create(classname=str(request.POST.get('fullname')), daylearn = str(start_date.date()), timelearnstart = time.timestart, timelearnend = time.timeend, dayname="Buổi " + str(count+1),active=0)
                        count = count+1
                start_date += step
            # return redirect('createclass')
            return redirect('listclass')
    context = {'form':form, 'unit':unit, 'area':area, 'room':room, 'timeshift':timeshift, 'timeweek':timeweek, 'teacher':teacher}
    return render(request, 'class/createClass.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff'])
def detailClass(request,pk):
    classname = Classname.objects.get(pk=pk)
    unit = Unit.objects.all()
    area = Area.objects.all()
    room = Room.objects.all()
    timeshift = TimeShift.objects.all()
    timeweek = TimeWeek. objects.all()
    teacher = Teacher.objects.all()
    form = UpdateClassForm(instance=classname)
    if request.method == 'POST':
        form = UpdateClassForm(request.POST, instance=classname)
        if form.is_valid():
            form.save()
            return redirect('listclass')
    context={'form':form, 'classname':classname, 'unit':unit, 'area':area, 'room':room, 'timeshift':timeshift, 'timeweek':timeweek, 'teacher':teacher}
    return render(request,'class/detailClass.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff'])
def detailClassStudent(request,pk):
    pk = pk
    classname = Classname.objects.get(pk=pk)
    liststudentinclass = Student.objects.all()
    timeshift = TimeShift.objects.get(pk=classname.timeshift_id)
    print(timeshift.timestart)
    schedule = Schedule.objects.all()
    context={'timeshift':timeshift,'liststudentinclass':liststudentinclass,'pk':pk, 'classname':classname, 'schedule':schedule}
    return render(request,'class/detailClassStudent.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff'])
def checkinClassStudent(request,pk):
    pk = pk
    classname = Classname.objects.get(pk=pk)
    liststudentinclass = Student.objects.all()
    timeshift = TimeShift.objects.get(pk=classname.timeshift_id)
    print(timeshift.timestart)
    checkinclass = CheckInClass.objects.all()
    if request.method=='POST':
        
        checkin = request.POST.getlist('check')
        for id in checkin:
            print(int(id))
            check = CheckInClass.objects.get(pk=int(id))
            if check.active == 0:
                check.active = 1
            check.save() 

        return redirect('listclass')
    schedule = Schedule.objects.all()
    export_movies_to_xlsx(request)
    context={'checkinclass':checkinclass,'timeshift':timeshift,'liststudentinclass':liststudentinclass,'pk':pk, 'classname':classname, 'schedule':schedule}
    return render(request,'class/checkinClassStudent.html',context)

# Student 
# Dat
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff'])
def listStudent(request):
    form = CreateStudentForm()
    liststudent = Student.objects.all()
    listclass = Classname.objects.all()
    export_movies_to_xlsx(request)
    for i1 in listclass:
        startdate = datetime.strptime(str(i1.startdate), '%Y-%m-%d')
        today = datetime.now()
        # student = Student.objects.all()
        if startdate <= today:
            # print(i1.pk) 
            i1.active=1
            i1.save()
        for a in liststudent:
            if i1.pk==a.classname_id:
                if startdate <= today:
                    # print(a.classname_id) 
                    a.active=1
                    a.save()
    context = {'listclass': listclass,'liststudent': liststudent}
    return render(request, 'student/student.html', context)

# @login_required(login_url='login')
# @allowed_users(allowed_roles=['admin','manager','staff'])
# def createDetailSchedule(request,pk):
#     student = Student.objects.get(pk=pk)
#     classrequest = Classname.objects.get(pk=student.classname_id)
#     time = TimeShift.objects.get(pk = classrequest.timeshift_id)
#     checkinclass = CheckInClass.objects.all()
#     if request.method == 'POST':
#         if student.havedetailschedule == 0:
#             student.havedetailschedule = 1
#             student.save()
#             count = 0
#             start_date = datetime.strptime(str(classrequest.startdate), '%Y-%m-%d')
#             step = timedelta(days=1)
#             # print(classrequest.timeweek_id)
#             if int(classrequest.timeweek_id) == 1:
#                 day = [0,2,4]
#             else:
#                 day = [1,3,5]
#             while count < int(classrequest.datecount):
#                 for i in day:
#                     if i==start_date.date().weekday():
#                         # print(start_date.date())
#                         CheckInClass.objects.create(id_student=pk,student=str(student.fullname),classname=str(classrequest.fullname), daylearn = str(start_date.date()), timelearnstart = time.timestart, timelearnend = time.timeend, dayname="Buổi " + str(count+1),active=0)
#                         count = count+1
#                 start_date += step
#             return redirect('liststudent')
#     context = {'checkinclass':checkinclass, 'classrequest':classrequest,'time':time,'student':student}
#     return render(request, 'student/createDetailSchedule.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff'])
def checkinStudent(request,pk):
    student = Student.objects.get(pk=pk) 
    classrequest = Classname.objects.get(pk=student.classname_id)
    time = TimeShift.objects.get(pk = classrequest.timeshift_id)
    checkinclass = CheckInClass.objects.filter(id_student=student.pk)
    context = {'checkinclass':checkinclass, 'classrequest':classrequest,'time':time,'student':student}
    if request.method == 'POST':
        if student.havedetailschedule == 0:
            student.havedetailschedule = 1
            student.save()
            # checkinclass = CheckInClass.objects.filter(id_student=student.pk) 
            count = 0
            start_date = datetime.strptime(str(classrequest.startdate), '%Y-%m-%d')
            step = timedelta(days=1)
            # print(classrequest.timeweek_id)
            if int(classrequest.timeweek_id) == 1:
                day = [0,2,4]
            else:
                day = [1,3,5]
            while count < int(classrequest.datecount):
                for i in day:
                    if i==start_date.date().weekday():
                        # print(start_date.date())
                        CheckInClass.objects.create(id_student=pk,student=str(student.fullname),classname=str(classrequest.fullname), daylearn = str(start_date.date()), timelearnstart = time.timestart, timelearnend = time.timeend, dayname="Buổi " + str(count+1),active=0)
                        count = count+1
                start_date += step
            # context = {'checkinclass':checkinclass, 'classrequest':classrequest,'time':time,'student':student}
            # return render(request, 'student/checkinStudent.html', context)
            return redirect('liststudent')
        else:
            student.havedetailschedule = 0
            student.active = 0
            student.save()
            CheckInClass.objects.filter(id_student=student.pk).delete()
            return redirect('liststudent')
    # context = {'checkinclass':checkinclass, 'classrequest':classrequest,'time':time,'student':student}
    return render(request, 'student/checkinStudent.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff'])
def createStudent(request):
    form = CreateStudentForm()
    gender = Gender.objects.all()
    unit = Unit.objects.all()
    classname = Classname.objects.all()
    if request.method == 'POST':
        form = CreateStudentForm(request.POST)
        if form.is_valid(): 
            form.save()
            send_mail(
                subject = 'XÁC NHẬN ĐĂNG KÝ HỌC VIÊN', # title mail
                message = 'Chào '+str(request.POST.get('fullname')) + ',\nBạn vừa hoàn thành đăng kí học viên tại HITECH. Trung tâm sẽ gửi thời khóa biểu cho bạn sớm nhất.', # nội dung mail
                from_email= None, # tài khoản
                auth_password= None, # mk
                recipient_list = [form.cleaned_data.get('email')],# mail người nhận
                fail_silently = False,
            )
            # return redirect('createstudent')
            return redirect('liststudent')
    context = { 'gender':gender, 'unit':unit, 'classname':classname}
    return render(request, 'student/createStudent.html', context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff'])
def detailStudent(request,pk):
    student = Student.objects.get(pk=pk)
    classname = Classname.objects.all()
    unit = Unit.objects.all()
    schedule = Schedule.objects.all()
    form = UpdateStudentForm(instance=student)
    if request.method == 'POST':
        form = UpdateStudentForm(request.POST, instance=student)
        if form.is_valid():
            # print(request.POST.get('classname'))
            student.classname_id = request.POST.get('classname')
            student.save()
            form.save()
            send_mail(
                subject = 'XÁC NHẬN ĐÃ THAY ĐỔI THÔNG TIN HỌC VIÊN', # title mail
                message = 'Chào '+ str(request.POST.get('fullname')) + ',\nThông tin học viên HITECH của bạn vừa được cập nhật. \nVui lòng trả lời tin mail này để xác nhận.', # nội dung mail
                from_email= None, # tài khoản
                auth_password= None, # mk
                recipient_list = [form.cleaned_data.get('email')],# mail người nhận
                fail_silently = False,
            )
            return redirect('liststudent')
    context = {'form':form, 'schedule':schedule, 'student':student, 'classname': classname, 'unit': unit}
    return render(request,'student/detailStudent.html',context)

# Teacher 
# Dat
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager','staff'])
def listTeacher(request):
    listteacher = Teacher.objects.all()
    array = []
    for item in listteacher:
        SL = {}
        SL['teachername'] = item.pk
        SL['number'] = Classname.objects.filter(teacher_id=item).count()
        array.append(SL)
        # print(array)
    context = {'listteacher': listteacher,'SL':array}
    return render(request, 'teacher/teacher.html', context)
    
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def createTeacher(request):
    gender = Gender.objects.all()
    listteacher = Teacher.objects.all()
    form = CreateTeacherForm()
    if request.method == 'POST':
        form = CreateTeacherForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listteacher')
    context = {'form':form, 'listteacher': listteacher, 'gender':gender}
    return render(request, 'teacher/createTeacher.html', context)

@login_required(login_url='login')
def detailTeacher(request,pk):
    teacher = Teacher.objects.get(pk=pk)
    gender = Gender.objects.all()
    form = UpdateTeacherForm(instance=teacher)
    if request.method == 'POST':
        form = UpdateTeacherForm(request.POST, instance=teacher)
        if form.is_valid():
            form.save()
            return redirect('listteacher')
    context = {'form':form, 'teacher':teacher, 'gender': gender}
    return render(request,'teacher/detailTeacher.html',context)

# Manager 
@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager'])
def listUser(request):
    listuser = User.objects.all()
    context={'listuser':listuser}
    return render(request,'manager/listUser.html',context)

@login_required(login_url='login')
def editProfile(request,pk):
    detailuser = User.objects.get(pk=pk)
    if request.method == 'POST':
        detailuser.last_name = request.POST.get('last_name')
        detailuser.first_name = request.POST.get('first_name')
        detailuser.save()
        return redirect('listuser')
    context={'detailuser':detailuser} 
    return render(request,'manager/editProfile.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def newRoom(request):
    form = CreateRoomForm()
    area = Area.objects.all()
    context={'form':form,'area':area} 
    if request.method == 'POST':
        form = CreateRoomForm(request.POST)
        if form.is_valid(): 
            form.save()
            return render(request,'manager/newRoom.html',context)
    return render(request,'manager/newRoom.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editRoom(request,pk):
    room = Room.objects.get(pk=pk)
    area = Area.objects.all()
    context={'area':area,'room':room} 
    if request.method == 'POST':
        room.fullname = request.POST.get('fullname')
        room.area_id = request.POST.get('area')
        if request.POST.get('active') == 'on':
            room.active = 1
        else:
            room.active = 0
        room.save()
        return redirect('/listarea/')
    return render(request,'manager/editRoom.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def newArea(request):
    form = CreateAreaForm()
    context={'form':form} 
    if request.method == 'POST':
        form = CreateAreaForm(request.POST)
        if form.is_valid(): 
            form.save()
            return redirect('newroom')
    return render(request,'manager/newArea.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editArea(request,pk):
    area = Area.objects.get(pk=pk)
    room = Room.objects.all()
    if request.method == 'POST':
        area.fullname = str(request.POST.get('fullname'))
        area.phonenumber = str(request.POST.get('phonenumber'))
        area.email = str(request.POST.get('email'))
        area.adress = str(request.POST.get('adress'))
        area.note = str(request.POST.get('note'))
        if request.POST.get('active') == 'on':
            area.active = 1
        else:
            area.active = 0
        area.save()
        return redirect('listarea')
    context={'area':area,'room':room}         
    return render(request,'manager/editArea.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def listArea(request):
    form = CreateAreaForm()
    listarea = Area.objects.all()
    context={'form':form,'listarea':listarea} 
    return render(request,'manager/listArea.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def notify(request):
    form = CreateNotifyForm()
    notify = Notify.objects.all()
    if request.method == 'POST':
        form = CreateNotifyForm(request.POST)
        if form.is_valid(): 
            form.save()
    context={'form':form,'notify':notify} 
    return render(request,'manager/notify.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editNotify(request,pk):
    noti = Notify.objects.get(pk=pk)
    if request.method == 'POST':
        if request.POST.get('active') == 'on':
            noti.active = 0
        else:
            noti.active = 1
        noti.content = request.POST.get('content')
        noti.save()
        print(noti.content)
        return redirect('notify')
    context={'noti':noti} 
    return render(request,'manager/editnotify.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def createTimeShift(request):
    form = CreateTimeShiftForm()
    timeshift = TimeShift.objects.all()
    if request.method == 'POST':
        form = CreateTimeShiftForm(request.POST)
        if form.is_valid(): 
            form.save()
    context={'form':form,'timeshift':timeshift} 
    return render(request,'manager/createTimeShift.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editTimeShift(request,pk):
    timeshift = TimeShift.objects.get(pk=pk)
    if request.method == 'POST':
        timeshift.fullname = request.POST.get('fullname')
        timeshift.infomation = request.POST.get('infomation')
        timeshift.timestart = request.POST.get('timestart')
        timeshift.timeend = request.POST.get('timeend')
        if request.POST.get('active') == 'on':
            timeshift.active = 1
        else:
            timeshift.active = 0
        timeshift.save()
        return redirect('createtimeshift')
    context={'timeshift':timeshift} 
    return render(request,'manager/editTimeShift.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def createUnit(request):
    form = CreateUnitForm()
    unit = Unit.objects.all()
    if request.method == 'POST':
        form = CreateUnitForm(request.POST)
        if form.is_valid(): 
            form.save()
    context={'form':form,'unit':unit} 
    return render(request,'manager/createUnit.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin'])
def editInformationUnit(request,pk):
    unit = Unit.objects.get(pk=pk)
    if request.method == 'POST':
        unit.infomation = request.POST.get('infomation')
        unit.fee = request.POST.get('fee')
        if request.POST.get('active') == 'on':
            unit.active = 1
        else:
            unit.active = 0
        unit.save()
        return redirect('createunit')
    context={'unit':unit} 
    return render(request,'manager/editInformationUnit.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager'])
def updateUser(request,pk):
    
    detailuser = User.objects.get(pk=pk)
    if request.method == 'POST':
        detailuser.last_name = request.POST.get('last_name')
        detailuser.first_name = request.POST.get('first_name')
        if str(request.POST.get('is_active')) == 'on':
            is_active = 1
        else:
            is_active = 0
        if str(request.POST.get('is_staff')) == 'on':
            is_staff = 1
        else:
            is_staff = 0
        if str(request.POST.get('is_superuser')) == 'on':
            is_superuser = 1
        else:
            is_superuser = 0
        print(is_active)
        print(is_staff)
        print(is_superuser)
        detailuser.is_active = is_active
        detailuser.is_staff = is_staff
        detailuser.is_superuser = is_superuser
        detailuser.groups.clear()
        detailuser.save()
        if request.POST.get('is_superuser'):
            group = Group.objects.get(name='admin')
            detailuser.groups.add(group)
        elif request.POST.get('is_staff'):
            group = Group.objects.get(name='manager')
            detailuser.groups.add(group)
        else:
            group = Group.objects.get(name='staff')
            detailuser.groups.add(group)
        return redirect('listuser')
    context={'detailuser':detailuser}
    return render(request,'manager/updateUser.html',context)

@login_required(login_url='login')
@allowed_users(allowed_roles=['admin','manager'])
def createStafAccount(request):
    form = CreateUserForm()
    context = {'form':form}
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            # user.is_superuser = 1
            # user.is_staff = 1
            user = form.save()  
            if str(request.POST.get('is_active')) == 'on':
                is_active = 1
            else:
                is_active = 0
            if str(request.POST.get('is_staff')) == 'on':
                is_staff = 1
            else:
                is_staff = 0
            if str(request.POST.get('is_superuser')) == 'on':
                is_superuser = 1
            else:
                is_superuser = 0
            username = form.cleaned_data.get('username')
            superuser = request.POST.get('is_staff')
            manager = request.POST.get('is_superuser')
            if is_superuser:
                group = Group.objects.get(name='admin')
                user.groups.add(group)
            elif is_staff:
                group = Group.objects.get(name='manager')
                user.groups.add(group)
            else:
                group = Group.objects.get(name='staff')
                user.groups.add(group)
            messages.success(request, 'Tài khoản ' + username + ' tạo thành công')
            return redirect('listuser') 
        else:
            messages.error(request, "Lỗi tạo tài khoản")
            return render(request,'manager/createStafAccount.html',context)
    return render(request,'manager/createStafAccount.html',context)

# Contact 
@login_required(login_url='login')
def contact(request):
    return render(request,'contact/contact.html')
 
# RoleError 
@login_required(login_url='login')
def errorRole(request):
    return render(request,'pages/errorrole.html')
  